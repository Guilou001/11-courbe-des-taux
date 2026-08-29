"""Trois sources libres : la courbe zéro-coupon de la BdC, les primes de terme Valet, les récessions C.D. Howe.

La courbe zéro-coupon (Bolder, Johnson et Metzler 2004) est publiée en composition continue
(BIS Papers 25, chapitre Canada) ; tout le dépôt garde cette convention : le taux forward entre
t1 et t2 vaut (z2*t2 - z1*t1)/(t2 - t1) et le facteur d'actualisation exp(-z*t).
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
import requests

RAW = Path("data/raw")

CURVE_URL = ("https://www.bankofcanada.ca/stats/results/csv"
             "?lookupPage=lookup_yield_curve.php&startRange=1986-01-01&searchRange=all")
TP_URL = "https://www.bankofcanada.ca/valet/observations/group/FVI_FINANCIAL_MARKETS_TP_GOC/json"
HOWE_URL = "https://cdhowe.org/wp-content/uploads/2024/06/Recession_Chronology_0.xlsx"

UA = {"User-Agent": "ycc laboratoire pedagogique (github.com/Guilou001/11-yield-curve-ca)"}


def fetch() -> None:
    """Télécharge les trois fichiers bruts (jamais commités)."""
    RAW.mkdir(parents=True, exist_ok=True)
    for url, name in [(CURVE_URL, "yield_curves.csv"), (TP_URL, "tp_goc.json"),
                      (HOWE_URL, "recession_chronology.xlsx")]:
        r = requests.get(url, headers=UA, timeout=300)
        r.raise_for_status()
        (RAW / name).write_bytes(r.content)


def load_curve() -> pd.DataFrame:
    """La courbe quotidienne : index dates, colonnes maturités en années (0,25 à 30), valeurs en %."""
    df = pd.read_csv(RAW / "yield_curves.csv", na_values=[" na", "na"])
    df.columns = [c.strip() for c in df.columns]
    df = df.drop(columns=[c for c in df.columns if not c], errors="ignore")
    df = df.loc[:, [c for c in df.columns if c == "Date" or c.startswith("ZC")]]
    df["Date"] = pd.to_datetime(df["Date"])
    df = df.set_index("Date").astype(float) * 100.0          # décimales -> %
    df.columns = [int(c[2:-2]) / 100.0 for c in df.columns]  # ZC1000YR -> 10.0 (années)
    return df.dropna(how="all")


def monthly_curve(curve: pd.DataFrame) -> pd.DataFrame:
    """La dernière courbe observée de chaque mois, index PeriodIndex mensuel."""
    m = curve.resample("ME").last().dropna(how="all")
    m.index = m.index.to_period("M")
    return m


def load_term_premium() -> pd.Series:
    """La prime de terme ACM à 10 ans (Valet, %), fin de mois, index PeriodIndex mensuel."""
    import json

    d = json.loads((RAW / "tp_goc.json").read_text())
    rows = [(o["d"], o.get("FVI_TP_GOC_10Y_ACM", {}).get("v")) for o in d["observations"]]
    s = pd.Series({pd.Timestamp(d_): float(v) for d_, v in rows if v not in (None, "")},
                  name="tp_10y").sort_index()
    s = s.resample("ME").last().dropna()
    s.index = s.index.to_period("M")
    return s


def load_recessions() -> list[tuple[pd.Period, pd.Period]]:
    """Les couples (pic, creux) mensuels de la chronologie du Business Cycle Council (2021, colonne révisée)."""
    from openpyxl import load_workbook

    ws = load_workbook(RAW / "recession_chronology.xlsx", read_only=True).active
    rows = [[c.value for c in row] for row in ws.iter_rows()]
    header_i, peak_j = None, None
    for i, row in enumerate(rows):
        for j, v in enumerate(row):
            if isinstance(v, str) and v.strip().startswith("Monthly Peak"):
                header_i, peak_j = i, j
    if header_i is None:
        raise ValueError("colonne « Monthly Peak » introuvable dans le xlsx C.D. Howe")

    def parse(v: str) -> pd.Period:
        return pd.Period(pd.Timestamp(v.split("(")[0].strip()), freq="M")

    out = []
    for row in rows[header_i + 1:]:
        peak, trough = row[peak_j], row[peak_j + 1]
        if isinstance(peak, str) and "(" in peak and isinstance(trough, str) and "(" in trough:
            out.append((parse(peak), parse(trough)))
    if len(out) < 10:
        raise ValueError(f"chronologie incomplète : {len(out)} récessions lues")
    return sorted(out)


def recession_indicator(index: pd.PeriodIndex, episodes: list[tuple[pd.Period, pd.Period]]) -> pd.Series:
    """1 pour un mois de récession : du mois SUIVANT le pic jusqu'au creux inclus (convention déclarée)."""
    flag = pd.Series(0, index=index, dtype=int)
    for peak, trough in episodes:
        flag[(index > peak) & (index <= trough)] = 1
    return flag


def recession_within(indicator: pd.Series, horizon: int = 12) -> pd.Series:
    """La cible du probit : 1 si au moins un mois de récession dans les `horizon` mois à venir.

    Les `horizon` derniers mois, dont l'avenir n'est pas observé, valent NaN.
    """
    future_any = (indicator[::-1].rolling(horizon).max()[::-1]).shift(-1)
    return future_any
