"""Le classeur de calculs obligataires : prix, duration et chocs IRRBB en formules Excel vivantes.

Le classeur reprend la convention du dépôt (composition continue) : le facteur d'actualisation
est EXP(-z*t) et le taux zéro de chaque flux vient de la feuille Courbe par RECHERCHEV exacte,
les flux tombant sur la grille de 0,25 an de la BdC.
"""

from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd

from ycc.alm import irrbb_shocks


def build_workbook(curve_row: pd.Series, curve_date: str, dest: Path,
                   coupon_pct: float = 4.0, maturity: float = 10.0) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    bold = Font(bold=True)

    ws0 = wb.active
    ws0.title = "Lisez-moi"
    lignes = [
        "Classeur de calculs obligataires sur la courbe zéro-coupon de la Banque du Canada.",
        f"Courbe du {curve_date} (source : bankofcanada.ca, décimales converties en %).",
        "Convention : composition continue, facteur d'actualisation EXP(-z*t) (BIS Papers 25).",
        "Feuille Obligation : changez le coupon (B2) ou l'échéance (B3), tout se recalcule.",
        "Feuille Chocs : les six scénarios IRRBB (CAD : 200/250/200 pb, BCBS 2024) revalorisent l'obligation.",
        "Les formules sont vivantes : aucun chiffre collé, tout part de la feuille Courbe.",
    ]
    for i, ligne in enumerate(lignes, start=1):
        ws0.cell(row=i, column=1, value=ligne)
    ws0.column_dimensions["A"].width = 100

    ws1 = wb.create_sheet("Courbe")
    ws1.cell(row=1, column=1, value="Échéance (années)").font = bold
    ws1.cell(row=1, column=2, value="Taux zéro (%)").font = bold
    for i, (tau, z) in enumerate(curve_row.items(), start=2):
        ws1.cell(row=i, column=1, value=float(tau))
        ws1.cell(row=i, column=2, value=round(float(z), 4))

    ws2 = wb.create_sheet("Obligation")
    ws2.cell(row=1, column=1, value="Paramètres").font = bold
    ws2.cell(row=2, column=1, value="Coupon annuel (%)")
    ws2.cell(row=2, column=2, value=coupon_pct)
    ws2.cell(row=3, column=1, value="Échéance (années)")
    ws2.cell(row=3, column=2, value=maturity)
    ws2.cell(row=4, column=1, value="Nominal")
    ws2.cell(row=4, column=2, value=100)
    ws2.cell(row=5, column=1, value="Coupons par an")
    ws2.cell(row=5, column=2, value=2)
    n_max = 60          # rangées jusqu'à 30 ans : l'échéance en B3 peut monter jusque-là
    head = ["t (années)", "Flux", "Taux zéro (%)", "Facteur EXP(-z*t)", "Flux actualisé", "t x flux actualisé"]
    for j, h in enumerate(head, start=1):
        ws2.cell(row=7, column=j, value=h).font = bold
    for k in range(1, n_max + 1):
        r = 7 + k
        ws2.cell(row=r, column=1, value=f"={k}/$B$5")
        ws2.cell(row=r, column=2,
                 value=f'=IF(A{r}>$B$3,0,$B$4*$B$2/100/$B$5+IF(A{r}=$B$3,$B$4,0))')
        ws2.cell(row=r, column=3, value=f"=VLOOKUP(A{r},Courbe!$A$2:$B$121,2,FALSE)")
        ws2.cell(row=r, column=4, value=f"=EXP(-C{r}/100*A{r})")
        ws2.cell(row=r, column=5, value=f"=B{r}*D{r}")
        ws2.cell(row=r, column=6, value=f"=A{r}*E{r}")
    last = 7 + n_max
    ws2.cell(row=last + 2, column=1, value="Prix").font = bold
    ws2.cell(row=last + 2, column=2, value=f"=SUM(E8:E{last})")
    ws2.cell(row=last + 3, column=1, value="Duration de Macaulay (années)").font = bold
    ws2.cell(row=last + 3, column=2, value=f"=SUM(F8:F{last})/B{last + 2}")
    for col, w in zip("ABCDEF", [22, 12, 14, 18, 16, 18], strict=True):
        ws2.column_dimensions[col].width = w

    ws3 = wb.create_sheet("Chocs")
    ws3.cell(row=1, column=1, value="Échéance (années)").font = bold
    ws3.cell(row=1, column=2, value="Courbe de base (%)").font = bold
    taus = np.array(curve_row.index, dtype=float)
    shocks = irrbb_shocks(taus)
    for j, name in enumerate(shocks, start=3):
        ws3.cell(row=1, column=j, value=name).font = bold
    for i, tau in enumerate(taus, start=2):
        ws3.cell(row=i, column=1, value=float(tau))
        ws3.cell(row=i, column=2, value=f"=Courbe!B{i}")
        for j, name in enumerate(shocks, start=3):
            ws3.cell(row=i, column=j, value=f"=$B{i}+{round(float(shocks[name][i - 2]), 4)}")
    # bloc d'aide : un rang par flux de l'obligation, taux choqué par RECHERCHEV exacte sur la grille
    hs = len(taus) + 4
    ws3.cell(row=hs - 1, column=1, value="t (années)").font = bold
    ws3.cell(row=hs - 1, column=2, value="Flux").font = bold
    for j, name in enumerate(shocks, start=3):
        ws3.cell(row=hs - 1, column=j, value=name).font = bold
    for k in range(n_max):
        r = hs + k
        ws3.cell(row=r, column=1, value=f"=Obligation!A{8 + k}")
        ws3.cell(row=r, column=2, value=f"=Obligation!B{8 + k}")
        for j in range(3, 3 + len(shocks)):
            ws3.cell(row=r, column=j,
                     value=f"=VLOOKUP($A{r},$A$2:$H${len(taus) + 1},{j},FALSE)")
    he = hs + n_max - 1
    r0 = he + 2
    ws3.cell(row=r0, column=1, value="Prix sous chaque scénario").font = bold
    ws3.cell(row=r0, column=2, value=f"=Obligation!B{last + 2}")
    ws3.cell(row=r0 - 1, column=2, value="base").font = bold
    for j, name in enumerate(shocks, start=3):
        col = get_column_letter(j)
        ws3.cell(row=r0 - 1, column=j, value=name).font = bold
        ws3.cell(row=r0, column=j,
                 value=f"=SUMPRODUCT($B${hs}:$B${he},EXP(-{col}{hs}:{col}{he}/100*$A${hs}:$A${he}))")
    ws3.column_dimensions["A"].width = 24

    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)
